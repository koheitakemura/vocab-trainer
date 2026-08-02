# -*- coding: utf-8 -*-
"""コース A（en-10-30k）の「簡単すぎる語」を多指標で洗い出す（判断ログ#34・層1+層2）。

判断ログ#32 は zipf>=4.3 の455候補をAIレビューして193語に絞ったが、閾値ゲートで
recall が頭打ちになっていた（zipf<4.3 の簡単語は一度も見られていない）。本スクリプトは
zipf という単一軸をやめ、**直交する4指標の和集合**で機械除去し、残りに容疑スコアを
振って AI レビュー対象を絞る。Kohei判断で「収録語数が減ってもよい」＝precision優先から
recall優先へ既定を反転している。

■ 機械除去（無条件・F1/F2/F4/F5 の和集合）
  F1 固有名詞: Tatoeba英文コーパスでの「非文頭かつ大文字始まり」の比率 >= 0.85。
      文頭は必ず大文字化されるので判定に使えない（除外している）。
      **JESCは使ってはいけない**——全文小文字化されており、足すと判定可能語は
      12,686→17,310に増える一方で比率が壊れ、検出が1,419→217語に激減する（実測）。
  F2 幼少期習得: AoA <= 8.0歳。carrot/spoon/puppy のような「幼児期に覚えるが大人の
      コーパスでは低頻度」＝zipfが原理的に分離できない語を落とす唯一の手段。
      **8.0を超えて緩めてはいけない**——<=10 まで緩めると behalf/legacy/heritage/
      constitutional/sacrifice のような10k帯で学ぶ価値のある語を巻き込む（実測）。
  F4 透明な派生: 既知語底面(NGSL/NAWL)からの -ly/-ness/-ful/-less/-wide 派生。
      意味が完全に予測可能な接尾辞だけに限る（-er/-tion/-ment 等は developer/
      completion のように独立した学習価値があるので対象外）。
  F5 辞書非収録: AoA(51,715語)・prevalence(61,853語)・EJDict-hand のどれにも
      載らない語。lmao/btw/gop/wtf/nyc のような略語・ネットスラング・固有名詞。
  F6 pos=固有名詞: 5.4のLLMコンテンツバッチが既に付けていた品詞タグ。判断ログ#30 は
      pos=数詞だけを機械除外の対象にしたが、固有名詞タグは使われないまま残っていた
      （wale「ウェールズ」・yorkshire・minneapolis・reuters 等が実際に生き残っていた）。
      F1（大文字率）はTatoebaでの出現が5回未満だと判定できないため、この品詞タグが
      その穴を埋める。判断ログ#30 が除外しないと決めた pos=代名詞/前置詞/接続詞
      （thou/thy/whomever/notwithstanding/albeit）は古語・文語であって基礎語ではない
      ので、ここでも対象にしない。

■ 使わないと決めた指標（実測に基づく不採用・再導入しないこと）
  prevalence 単独: >=2.3 で3,529語を捕まえるが、中身が ambiguous/cynical/endorse/
      compassion/treason/excessive/testimony ——日本人学習者が10k帯で学ぶ価値のある語
      を大量に巻き込む。prevalence も AoA も「ネイティブの習得」を測る指標であって
      L2学習者の難易度とは別軸（#32でzipfが失敗したのと同じ罠の一段深い版）。
      AoAが例外的に使えるのは <=8歳＝L2学習者でも確実に既習の領域を切る場合だけ。
      よって prevalence は容疑スコアの加点とF5の実在レンマ判定にのみ使う。
  米国姓リスト単独: 上位1,797姓で295語ヒットするが angel/archer/bacon/baker/berry/
      bishop/cherry/fox/frost/grace 等の普通名詞が大半。容疑スコア専用。

入力:
  raw/norms/*                        （fetch_difficulty_norms.py の出力）
  raw/eng_sentences_full.tsv.bz2     （Tatoeba・F1の大文字率算出用）
  raw/en-wordfreq-ranked.json        （zipf）
  raw/ngsl-1.2-lemmatized.csv, raw/nawl-1.2-lemmatized.csv
  ../public/data/courses/en-10-30k/  （現行カード）

出力:
  raw/en-easy-auto-remove.json  機械除去の確定リスト
  raw/en-easy-gray-band.json    AIレビュー対象（容疑スコア降順）

実行: python analyze_easy_words_en.py
"""
import bz2
import json
import os
import re
import sys
from collections import defaultdict

import ejdict

RAW = "raw"
NORMS = f"{RAW}/norms"
COURSE = "../public/data/courses/en-10-30k"
CAP_CACHE = f"{NORMS}/en_cap_counts.json"

AOA_MAX = 8.0            # F2。緩めるな（上記コメント参照）
CAP_MIN_RATIO = 0.85     # F1
CAP_MIN_OBS = 5          # F1: これ未満の出現数では比率が信用できない
SAFE_SUFFIXES = ("ly", "ness", "ful", "less", "wide")  # F4
GRAY_MIN_SCORE = 2       # 容疑スコアがこれ以上なら AI レビューへ
TOKEN_RE = re.compile(r"[A-Za-z]+")


def log(msg: str) -> None:
    print(f"[analyze_easy_words_en] {msg}", flush=True)


# ---------- 規範データ ----------
def load_xlsx_column(path: str, value_columns: tuple[str, ...]) -> dict[str, float]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    col = next((header.index(c) for c in value_columns if c in header), None)
    if col is None:
        raise SystemExit(f"{path}: 想定した列 {value_columns} が見つからない（header={header[:12]}）")
    out: dict[str, float] = {}
    for r in rows:
        if not r or r[0] is None or r[col] is None:
            continue
        try:
            out[str(r[0]).strip().lower()] = float(r[col])
        except (TypeError, ValueError):
            continue
    log(f"{os.path.basename(path)}[{header[col]}]: {len(out)}語")
    return out


def load_known_baseline() -> set[str]:
    known: set[str] = set()
    for name in ("ngsl-1.2-lemmatized.csv", "nawl-1.2-lemmatized.csv"):
        with open(f"{RAW}/{name}", encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("##"):
                    continue
                known.update(w.strip().lower() for w in line.split(",") if w.strip())
    log(f"既知語底面(NGSL+NAWL 全活用形): {len(known)}語")
    return known


# ---------- F1: Tatoeba での非文頭大文字率 ----------
def load_capitalization(targets: set[str]) -> dict[str, list[int]]:
    if os.path.exists(CAP_CACHE):
        cached = json.load(open(CAP_CACHE, encoding="utf-8"))
        if set(cached) >= targets:
            log(f"大文字率キャッシュを再利用: {CAP_CACHE}")
            return cached
    log("Tatoeba英文を走査して非文頭大文字率を算出中（JESCは小文字化済のため使わない）...")
    mid_total: dict[str, int] = defaultdict(int)
    mid_upper: dict[str, int] = defaultdict(int)
    n = 0
    with bz2.open(f"{RAW}/eng_sentences_full.tsv.bz2", "rt", encoding="utf-8") as f:
        for line in f:
            n += 1
            parts = line.rstrip("\n").split("\t")
            if len(parts) < 3:
                continue
            for i, tok in enumerate(TOKEN_RE.findall(parts[2])):
                if i == 0:
                    continue  # 文頭は必ず大文字化されるので判定に使えない
                low = tok.lower()
                if low in targets:
                    mid_total[low] += 1
                    if tok[0].isupper():
                        mid_upper[low] += 1
    counts = {w: [mid_upper.get(w, 0), mid_total.get(w, 0)] for w in targets}
    with open(CAP_CACHE, "w", encoding="utf-8") as f:
        json.dump(counts, f)
    judgeable = sum(1 for v in counts.values() if v[1] >= CAP_MIN_OBS)
    log(f"Tatoeba {n}行を走査。判定可能(出現{CAP_MIN_OBS}回以上): {judgeable}/{len(targets)}語")
    return counts


def cap_ratio(counts: dict[str, list[int]], word: str) -> float | None:
    up, total = counts.get(word, [0, 0])
    return up / total if total >= CAP_MIN_OBS else None


# ---------- main ----------
def main():
    aoa = load_xlsx_column(f"{NORMS}/AoA_51715_words.xlsx", ("AoA_Kup_lem", "Rating.Mean"))
    prevalence = load_xlsx_column(f"{NORMS}/English_Word_Prevalences.xlsx", ("Prevalence",))
    surnames = set(json.load(open(f"{NORMS}/us_surnames.json", encoding="utf-8")))
    log(f"米国上位姓: {len(surnames)}件（容疑スコア専用・自動削除には使わない）")
    known = load_known_baseline()
    ej_index, _ = ejdict.build_index()
    zipf = {d["lemma"]: d["zipf"] for d in json.load(open(f"{RAW}/en-wordfreq-ranked.json", encoding="utf-8"))}

    manifest = json.load(open(f"{COURSE}/manifest.json", encoding="utf-8"))
    cards = []
    for band in manifest["bands"]:
        cards += json.load(open(f"{COURSE}/{band}", encoding="utf-8"))
    by_headword = {c["headword"]: c for c in cards}
    log(f"現行カード: {len(cards)}件")

    counts = load_capitalization(set(by_headword))

    def f1_proper_noun(w: str) -> bool:
        r = cap_ratio(counts, w)
        return r is not None and r >= CAP_MIN_RATIO

    def f2_childhood(w: str) -> bool:
        return w in aoa and aoa[w] <= AOA_MAX

    def f4_transparent_derivative(w: str) -> bool:
        for suf in SAFE_SUFFIXES:
            if w.endswith(suf) and len(w) > len(suf) + 3:
                base = w[: -len(suf)]
                if base in known or base + "e" in known:
                    return True
        return False

    def f5_not_a_lexicon_word(w: str) -> bool:
        return w not in aoa and w not in prevalence and not ejdict.resolve_gloss(w, ej_index)

    def f6_pos_proper_noun(w: str) -> bool:
        return by_headword[w].get("pos") == "固有名詞"

    FILTERS = (
        ("F1_proper_noun", f1_proper_noun),
        ("F2_childhood_aoa", f2_childhood),
        ("F4_transparent_derivative", f4_transparent_derivative),
        ("F5_not_in_lexicon", f5_not_a_lexicon_word),
        ("F6_pos_proper_noun", f6_pos_proper_noun),
    )

    auto_remove: dict[str, list[str]] = {}
    for w in by_headword:
        hit = [name for name, fn in FILTERS if fn(w)]
        if hit:
            auto_remove[w] = hit

    log("=" * 70)
    for name, _ in FILTERS:
        log(f"  {name}: {sum(1 for v in auto_remove.values() if name in v)}語")
    log(f"機械除去(和集合): {len(auto_remove)}語 -> 残 {len(cards) - len(auto_remove)}語")

    # ---------- 層2: 残りに容疑スコアを振る ----------
    # 単独では削除根拠にならないが、重なると「簡単すぎる/語彙ではない」可能性が高い信号。
    def clipped_form(w: str) -> bool:
        """app<-application, info<-information のような短縮形。既知語の真部分接頭辞。"""
        if len(w) > 5:
            return False
        return any(k != w and k.startswith(w) and len(k) >= len(w) + 3 for k in known)

    gray = []
    for w, card in by_headword.items():
        if w in auto_remove:
            continue
        score = 0
        why = []
        if w in surnames:
            score += 3
            why.append("米国上位姓")
        r = cap_ratio(counts, w)
        if r is not None and 0.5 <= r < CAP_MIN_RATIO:
            score += 3
            why.append(f"大文字率{r:.2f}")
        elif r is not None and 0.3 <= r < 0.5:
            score += 1
            why.append(f"大文字率{r:.2f}")
        if clipped_form(w):
            score += 3
            why.append("短縮形の疑い")
        if w in aoa and aoa[w] <= 10.0:
            score += 2
            why.append(f"AoA{aoa[w]:.1f}歳")
        elif w in aoa and aoa[w] <= 12.0:
            score += 1
            why.append(f"AoA{aoa[w]:.1f}歳")
        if w in prevalence and prevalence[w] >= 2.3:
            score += 2
            why.append(f"prevalence{prevalence[w]:.2f}")
        elif w in prevalence and prevalence[w] >= 2.0:
            score += 1
            why.append(f"prevalence{prevalence[w]:.2f}")
        if zipf.get(w, 0) >= 4.0:
            # 固有名詞を先に落としてあるので、この段階の高頻度は素直に「よく使う＝易しい」を意味する
            score += 1
            why.append(f"zipf{zipf.get(w):.2f}")
        if not ejdict.resolve_gloss(w, ej_index):
            score += 1
            why.append("EJDict非収録")
        if score >= GRAY_MIN_SCORE:
            gray.append(
                {
                    "id": card["id"],
                    "headword": w,
                    "gloss": card.get("gloss"),
                    "pos": card.get("pos"),
                    "score": score,
                    "signals": why,
                }
            )
    gray.sort(key=lambda x: (-x["score"], x["headword"]))
    log(f"容疑スコア>={GRAY_MIN_SCORE} でAIレビュー対象: {len(gray)}語")
    for s in range(10, 1, -1):
        n = sum(1 for g in gray if g["score"] == s)
        if n:
            log(f"    score {s}: {n}語")

    with open(f"{RAW}/en-easy-auto-remove.json", "w", encoding="utf-8") as f:
        json.dump(
            {"count": len(auto_remove),
             "words": {w: {"id": by_headword[w]["id"], "filters": v} for w, v in sorted(auto_remove.items())}},
            f, ensure_ascii=False, indent=1)
    with open(f"{RAW}/en-easy-gray-band.json", "w", encoding="utf-8") as f:
        json.dump({"count": len(gray), "items": gray}, f, ensure_ascii=False, indent=1)
    log(f"-> {RAW}/en-easy-auto-remove.json / {RAW}/en-easy-gray-band.json")

    survivors = [w for w in by_headword if w not in auto_remove and all(g["headword"] != w for g in gray)]
    log(f"無条件keep（機械的に難語と確定）: {len(survivors)}語")
    log("Done.")


if __name__ == "__main__":
    sys.exit(main())
