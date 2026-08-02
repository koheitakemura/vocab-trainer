# -*- coding: utf-8 -*-
"""除去後の en-10-30k を検分する（判断ログ#34の効果測定）。

「簡単な語が減ったか」を主観でなく数字で確認するための後追いレポート。
除去前後で各シグナルの残存率を並べ、頻度上位の残存語を目視できる形で出す。

実行: python report_easy_words_en.py
"""
import json
import os
import sys

import ejdict

RAW = "raw"
NORMS = f"{RAW}/norms"
COURSE = "../public/data/courses/en-10-30k"


def log(msg: str) -> None:
    print(f"[report_easy_words_en] {msg}", flush=True)


def load_xlsx_column(path, cols):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    ci = next(header.index(c) for c in cols if c in header)
    out = {}
    for r in rows:
        if not r or r[0] is None or r[ci] is None:
            continue
        try:
            out[str(r[0]).strip().lower()] = float(r[ci])
        except (TypeError, ValueError):
            pass
    return out


def main():
    aoa = load_xlsx_column(f"{NORMS}/AoA_51715_words.xlsx", ("AoA_Kup_lem",))
    prevalence = load_xlsx_column(f"{NORMS}/English_Word_Prevalences.xlsx", ("Prevalence",))
    surnames = set(json.load(open(f"{NORMS}/us_surnames.json", encoding="utf-8")))
    cap = json.load(open(f"{NORMS}/en_cap_counts.json", encoding="utf-8"))
    ej_index, _ = ejdict.build_index()

    manifest = json.load(open(f"{COURSE}/manifest.json", encoding="utf-8"))
    cards = []
    for band in manifest["bands"]:
        cards += json.load(open(f"{COURSE}/{band}", encoding="utf-8"))
    hw = [c["headword"] for c in cards]
    log(f"現在の収録語数: {len(hw)}（manifest.wordCount={manifest['wordCount']}）")
    assert len(hw) == manifest["wordCount"], "manifest.wordCount と実カード数が不一致"

    removed_path = f"{RAW}/en-10-30k-easy-removed.json"
    if os.path.exists(removed_path):
        rm = json.load(open(removed_path, encoding="utf-8"))
        log(f"除去済み見出し語: {rm['count']}語")

    def ratio(w):
        u, t = cap.get(w, [0, 0])
        return u / t if t >= 5 else None

    checks = {
        "固有名詞疑い(大文字率>=0.85)": lambda w: (ratio(w) or 0) >= 0.85,
        "大文字率0.5-0.85のグレー": lambda w: 0.5 <= (ratio(w) or 0) < 0.85,
        "AoA<=8歳": lambda w: w in aoa and aoa[w] <= 8.0,
        "AoA 8-10歳": lambda w: w in aoa and 8.0 < aoa[w] <= 10.0,
        "prevalence>=2.3": lambda w: w in prevalence and prevalence[w] >= 2.3,
        "米国上位姓と同綴り": lambda w: w in surnames,
        "EJDict非収録": lambda w: not ejdict.resolve_gloss(w, ej_index),
    }
    log("=" * 60)
    log("残存語に各シグナルがどれだけ残っているか:")
    for name, fn in checks.items():
        n = sum(1 for w in hw if fn(w))
        log(f"  {name}: {n}語 ({n / len(hw) * 100:.1f}%)")

    aoa_vals = [aoa[w] for w in hw if w in aoa]
    if aoa_vals:
        aoa_vals.sort()
        log("=" * 60)
        log(f"残存語のAoA分布（{len(aoa_vals)}語でカバー）: "
            f"中央値 {aoa_vals[len(aoa_vals)//2]:.1f}歳 / "
            f"下位5% {aoa_vals[len(aoa_vals)//20]:.1f}歳 / 最小 {aoa_vals[0]:.1f}歳")

    log("=" * 60)
    log("残存語の頻度上位80語（ここが簡単に見えるならまだ緩い）:")
    ordered = sorted(cards, key=lambda c: c["frequencyRank"])
    print("  " + ", ".join(c["headword"] for c in ordered[:80]))
    log("残存語で最もAoAが低い40語:")
    low = sorted(((aoa[w], w) for w in hw if w in aoa))[:40]
    print("  " + ", ".join(f"{w}({a:.1f})" for a, w in low))
    log("Done.")


if __name__ == "__main__":
    sys.exit(main())
